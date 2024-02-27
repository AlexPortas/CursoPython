import openpyxl

# Abrir un archivo Excel (workbook)
wb = openpyxl.load_workbook("02_Excel_data.xlsx")

# Imprimir los nombres de las hojas
print("Nombres de hojas: ")
print(wb.sheetnames)

print("\nNombres de hojas (otra forma): ")
for sheet in wb:
    print(sheet.title)

# Crear una variable que haga referencia a la primera hoja del excel
hoja_uno = wb.sheetnames[0]
print("\nPrimera hoja: ")
print(hoja_uno)

# Crear una variable que haga referencia a la primera hoja del excel
hoja_uno = wb["address"]

# Acceder a una celda directamente
print(hoja_uno["A1"].value) # Contenido
print(hoja_uno["B1"]) # Referencia

# Acceder a una celda a través de una nomenclatura fila-columna
print(hoja_uno.cell(row=5,column=2).value)

# Acceso a múltiples celdas de una hoja concreta

multiple_cells = hoja_uno["A1":"B3"]
for row in multiple_cells:
    for cell in row:
        print(cell.value)

# Acceso a todas las filas de una hoja concreta

for fila in hoja_uno.rows:
        for columna in fila:
            print(columna.coordinate, columna.value)

# Acceso a todas las columnas de una hoja concreta

for columna in hoja_uno.columns:
    for fila in columna:
        print(fila.coordinate, fila.value)

# Modificación de celdas de una hoja concreta

import datetime

# Crear una variable que haga referencia a la segunda hoja del excel
hoja_dos = wb["ventas"]

# Modificar una celda (3 formas diferentes)
hoja_dos["B6"] = 5
hoja_dos["C6"].value = 5
hoja_dos.cell(row=6, column=4, value=5)

# Añadir la hora actual en la celda E2
hoja_dos["E2"].value = datetime.datetime.now()

# Guardar cambios (IMPORTANTE! El excel debe estar cerrado)
wb.save("02_Excel_data.xlsx")

# Modificar la celda B2, que contiene un 40, por un 99
print("Celda B2 antes de modificarla: ", hoja_dos["B2"].value)
hoja_dos["B2"] = 99
print("Celda B2 despues de modificarla: ", hoja_dos["B2"].value)

# Modificar la celda C2, pero esta vez con una formula
print("Celda C2 antes de modificarla con la formula: ", hoja_dos["C2"].value)
hoja_dos["C2"] = "=SUM(B2, 3)"
print("Celda C2 despues de modificarla con la formula: ", hoja_dos["C2"].value)

# Guardar cambios
wb.save("02_Excel_data.xlsx")

## Añadir contenido nuevo a una hoja concreta

hoja_dos.append(["Junio",99,100])
# Guardar cambios
wb.save("02_Excel_data.xlsx")

## Crear una hoja nueva y darle nombre

# Abrir un archivo Excel (workbook)
wb = openpyxl.load_workbook("02_Excel_data.xlsx")

# Creamos una nueva hoja (por defecto al final, si no, le podemos pasar como parametro el indice de la posicion que se quiera)
hoja_otros = wb.create_sheet("Mysheet")

# Nombre de la hoja
hoja_otros.title = "Otros"

# Imprimir los nombres de las hojas
print("Nombres de hojas: ")
print(wb.sheetnames)

# Guardar cambios
wb.save("02_Excel_data.xlsx")